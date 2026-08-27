from flask import Flask, render_template, request, jsonify, Response, session
from werkzeug.security import generate_password_hash, check_password_hash
import database
import json
import csv
import io
import os
import ssl

from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'proyectabm_bmining_secret_key_2026')
app.permanent_session_lifetime = timedelta(minutes=5)

ALLOWED_DOMAIN = "@bmining.cl"

def is_valid_bmining_email(email):
    return bool(email and email.lower().strip().endswith(ALLOWED_DOMAIN))

def get_current_user_info():
    user_id = session.get('user_id')
    if user_id:
        user = database.get_user_by_id(user_id)
        if user:
            return user.get('email', 'Anónimo'), user.get('nombre', 'Usuario')
    return 'Sistema', 'Invitado'

# Initialize database tables and seed data
database.init_db()

@app.route('/')
def index():
    session.pop('user_id', None)
    return render_template('index.html')

# AUTHENTICATION ROUTES (@bmining.cl ONLY)
@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({"status": "unauthenticated"}), 200
    user = database.get_user_by_id(user_id)
    if not user:
        session.pop('user_id', None)
        return jsonify({"status": "unauthenticated"}), 200
    return jsonify({"status": "success", "user": user})

@app.route('/api/auth/register', methods=['POST'])
def api_auth_register():
    data = request.get_json() or {}
    nombre = data.get('nombre', '').strip()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not nombre or not email or not password:
        return jsonify({"status": "error", "message": "Todos los campos son obligatorios"}), 400

    if not is_valid_bmining_email(email):
        return jsonify({
            "status": "error",
            "message": f"Acceso restringido: El correo debe ser institucional con dominio {ALLOWED_DOMAIN}"
        }), 400

    if len(password) < 6:
        return jsonify({"status": "error", "message": "La contraseña debe tener al menos 6 caracteres"}), 400

    existing_user = database.get_user_by_email(email)
    if existing_user:
        return jsonify({"status": "error", "message": "El correo ya se encuentra registrado. Inicia sesión."}), 400

    try:
        pw_hash = generate_password_hash(password)
        user_id = database.create_user(nombre, email, pw_hash)
        session['user_id'] = user_id
        user = database.get_user_by_id(user_id)
        return jsonify({"status": "success", "message": "Registro exitoso", "user": user})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error en el registro: {str(e)}"}), 500

@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({"status": "error", "message": "Ingresa correo y contraseña"}), 400

    if not is_valid_bmining_email(email):
        return jsonify({
            "status": "error",
            "message": f"Acceso restringido: Solo se permiten correos institucionales con dominio {ALLOWED_DOMAIN}"
        }), 400

    user = database.get_user_by_email(email)
    if not user:
        name_part = email.split('@')[0].replace('.', ' ').title()
        pw_hash = generate_password_hash(password or 'bmining_pass_123')
        user_id = database.create_user(name_part, email, pw_hash)
        user = database.get_user_by_id(user_id)

    session.permanent = True
    session['user_id'] = user['id']
    user_info = {"id": user['id'], "nombre": user['nombre'], "email": user['email']}
    database.log_activity(user['email'], user['nombre'], 'INICIO_SESION', 'Inicio de sesión exitoso', request.remote_addr)
    return jsonify({"status": "success", "message": "Inicio de sesión exitoso", "user": user_info})

@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    user_id = session.get('user_id')
    if user_id:
        u = database.get_user_by_id(user_id)
        if u:
            database.log_activity(u['email'], u['nombre'], 'CIERRE_SESION', 'Cierre de sesión manual', request.remote_addr)
    session.pop('user_id', None)
    return jsonify({"status": "success", "message": "Sesión cerrada correctamente"})

@app.route('/api/logs', methods=['GET'])
def api_get_logs():
    logs = database.get_audit_logs(limit=200)
    return jsonify({"status": "success", "data": logs})

@app.route('/api/logs/export', methods=['GET'])
def api_export_logs():
    fmt = request.args.get('format', 'csv').lower()
    logs = database.get_audit_logs(limit=1000)
    
    if fmt == 'txt':
        output_lines = []
        output_lines.append("================================================================================")
        output_lines.append("                 HOURCAST / PROYECTABM - REGISTRO DE AUDITORÍA                  ")
        output_lines.append(f" Reporte generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        output_lines.append("================================================================================\n")
        
        for log in logs:
            fecha = str(log.get('fecha_registro', ''))[:19]
            nombre = log.get('usuario_nombre') or 'Usuario'
            email = log.get('usuario_email') or 'N/A'
            accion = log.get('accion') or 'ACCION'
            detalles = log.get('detalles') or ''
            ip = log.get('ip_address') or '127.0.0.1'
            output_lines.append(f"[{fecha}] [{accion}] {nombre} ({email}) [IP: {ip}]\n  Detalles: {detalles}\n")
            
        content = "\n".join(output_lines)
        return Response(
            content.encode('utf-8'),
            mimetype='text/plain',
            headers={"Content-Disposition": "attachment; filename=historial_auditoria_logs.txt"}
        )
    else:
        # Default CSV Export (delimiter = ';')
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(['ID', 'Fecha_Hora', 'Usuario_Nombre', 'Usuario_Email', 'Accion', 'Detalles', 'IP_Address'])
        for log in logs:
            fecha = str(log.get('fecha_registro', ''))[:19]
            writer.writerow([
                log.get('id', ''),
                fecha,
                log.get('usuario_nombre', ''),
                log.get('usuario_email', ''),
                log.get('accion', ''),
                log.get('detalles', ''),
                log.get('ip_address', '')
            ])
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={"Content-Disposition": "attachment; filename=historial_auditoria_logs.csv"}
        )

# Endpoint to fetch live UF from mindicador.cl/api
@app.route('/api/uf', methods=['GET'])
def api_get_uf():
    # Provider 1: mindicador.cl/api
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(
            'https://mindicador.cl/api',
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)', 'Accept': 'application/json'}
        )
        with urllib.request.urlopen(req, timeout=4, context=ctx) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            if 'uf' in res_data and 'valor' in res_data['uf']:
                uf_val = float(res_data['uf']['valor'])
                uf_fecha = str(res_data['uf']['fecha'])[:10]
                return jsonify({"status": "success", "valor": uf_val, "fecha": uf_fecha, "origen": "mindicador.cl"})
    except Exception as e1:
        pass

    # Provider 2: mindicador.cl/api/uf
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(
            'https://mindicador.cl/api/uf',
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)', 'Accept': 'application/json'}
        )
        with urllib.request.urlopen(req, timeout=4, context=ctx) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            if 'serie' in res_data and len(res_data['serie']) > 0:
                latest = res_data['serie'][0]
                uf_val = float(latest['valor'])
                uf_fecha = str(latest['fecha'])[:10]
                return jsonify({"status": "success", "valor": uf_val, "fecha": uf_fecha, "origen": "mindicador.cl/uf"})
    except Exception as e2:
        pass

    return jsonify({"status": "success", "valor": 38850.00, "fecha": "Hoy", "origen": "UF Base"})

# API Endpoints for Perfiles
@app.route('/api/perfiles', methods=['GET'])
def api_get_perfiles():
    perfiles = database.get_perfiles()
    return jsonify({"status": "success", "data": perfiles})

@app.route('/api/perfiles', methods=['POST'])
def api_add_perfil():
    data = request.get_json() or {}
    nombre = data.get('nombre', '').strip()
    tarifa = data.get('tarifa_costo', 0)
    descripcion = data.get('descripcion', '').strip()
    
    if not nombre or float(tarifa) <= 0:
        return jsonify({"status": "error", "message": "Nombre y tarifa de costo válidos son obligatorios"}), 400
        
    try:
        new_id = database.add_perfil(nombre, tarifa, descripcion)
        u_email, u_nombre = get_current_user_info()
        database.log_activity(u_email, u_nombre, 'CREAR_PERFIL', f'Registró perfil "{nombre}" con tarifa {tarifa} UF/h', request.remote_addr)
        return jsonify({"status": "success", "id": new_id, "message": "Perfil registrado exitosamente"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error al guardar perfil: {str(e)}"}), 500

@app.route('/api/perfiles/<int:perfil_id>', methods=['DELETE'])
def api_delete_perfil(perfil_id):
    try:
        database.delete_perfil(perfil_id)
        return jsonify({"status": "success", "message": "Perfil eliminado"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

import openpyxl
import openpyxl.styles
import openpyxl.utils

@app.route('/api/perfiles/template_excel', methods=['GET'])
def api_download_template_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga Perfiles"

    headers = ["Profesional", "Perfil", "Precio Costo"]
    ws.append(headers)

    green_fill = openpyxl.styles.PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    font_bold = openpyxl.styles.Font(bold=True, color="000000")
    border_thin = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = green_fill
        cell.font = font_bold
        cell.border = border_thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    sample_data = [
        ["Rodrigo Poblete", "Consultor Senior", 1.8],
        ["Catalina Olivares", "Ingeniero Especialista", 0.7],
        ["Juan Pedreros", "Ingeniero Senior", 1.1],
        ["Amilcar Chavez", "Ingeniero", 0.5]
    ]

    for row_data in sample_data:
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
        for cell in row:
            cell.border = border_thin

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Plantilla_Carga_Masiva_ProyectaBM.xlsx"}
    )

@app.route('/api/perfiles/upload_excel', methods=['POST'])
def api_upload_excel_perfiles():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No se adjuntó ningún archivo Excel"}), 400

    file = request.files['file']
    filename = file.filename.lower()

    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({"status": "error", "message": "Formato no válido. Debe ser un archivo Excel (.xlsx, .xls) o CSV (.csv)"}), 400

    # Clear previous database profiles for bulk upload replacement
    database.clear_all_perfiles()

    imported_items = []
    try:
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            csv_reader = csv.reader(stream)
            header = next(csv_reader, None)
            for row in csv_reader:
                if not row or len(row) < 3:
                    continue
                profesional = row[0].strip()
                perfil_nombre = row[1].strip()
                tarifa_raw = str(row[2]).strip().replace(',', '.')
                try:
                    tarifa_costo = float(tarifa_raw)
                except ValueError:
                    continue
                if perfil_nombre and tarifa_costo > 0:
                    pid = database.upsert_perfil(perfil_nombre, tarifa_costo)
                    imported_items.append({
                        "profesional": profesional or perfil_nombre,
                        "perfil_id": pid,
                        "perfil_nombre": perfil_nombre,
                        "tarifa_costo": tarifa_costo
                    })
        else:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return jsonify({"status": "error", "message": "El archivo Excel está vacío"}), 400

            start_row = 1 if isinstance(rows[0][0], str) and ("profesional" in str(rows[0][0]).lower() or "nombre" in str(rows[0][0]).lower()) else 0

            for row in rows[start_row:]:
                if not row or len(row) < 3:
                    continue
                profesional = str(row[0]).strip() if row[0] is not None else ""
                perfil_nombre = str(row[1]).strip() if row[1] is not None else ""
                tarifa_raw = str(row[2]).strip().replace(',', '.') if row[2] is not None else "0"
                try:
                    tarifa_costo = float(tarifa_raw)
                except ValueError:
                    continue

                if perfil_nombre and tarifa_costo > 0:
                    pid = database.upsert_perfil(perfil_nombre, tarifa_costo)
                    imported_items.append({
                        "profesional": profesional or perfil_nombre,
                        "perfil_id": pid,
                        "perfil_nombre": perfil_nombre,
                        "tarifa_costo": tarifa_costo
                    })

        u_email, u_nombre = get_current_user_info()
        database.log_activity(u_email, u_nombre, 'CARGA_MASIVA_EXCEL', f'Procesó archivo Excel e importó {len(imported_items)} registros', request.remote_addr)

        # Auto save projection entry so all users see the 45 items in shared projections
        saved_proj_id = None
        if imported_items:
            try:
                # Format full items list for projection engine
                formatted_items = []
                for idx, imp in enumerate(imported_items):
                    formatted_items.append({
                        "uid": f"imp_{idx+1}_{int(datetime.now().timestamp())}",
                        "profesional": imp.get("profesional") or f"Profesional {idx+1}",
                        "perfil_id": imp.get("perfil_id", ""),
                        "perfil_nombre": imp.get("perfil_nombre", "Personalizado"),
                        "tarifa_costo": float(imp.get("tarifa_costo", 1.0)),
                        "horas": 10,
                        "costo_total": 0,
                        "margen_porcentaje": 15.0,
                        "monto_utilidad": 0,
                        "precio_venta": 0
                    })
                saved_proj_id = database.save_proyeccion(
                    nombre_proyecto=f"Carga Masiva Excel - {u_nombre}",
                    cliente="Bmining",
                    modo_margen="costo",
                    margen_global=15.0,
                    unidad_escala=1000.0,
                    items_data=formatted_items
                )
            except Exception as ex_proj:
                pass

        return jsonify({
            "status": "success",
            "message": f"Se importaron {len(imported_items)} registros exitosamente",
            "count": len(imported_items),
            "items": imported_items,
            "projection_id": saved_proj_id
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error al procesar archivo Excel: {str(e)}"}), 500

# API Endpoints for Proyecciones
@app.route('/api/proyecciones', methods=['GET'])
def api_get_proyecciones():
    proyecciones = database.get_proyecciones()
    return jsonify({"status": "success", "data": proyecciones})

@app.route('/api/proyecciones/<int:proyeccion_id>', methods=['GET'])
def api_get_proyeccion(proyeccion_id):
    proyeccion = database.get_proyeccion(proyeccion_id)
    if not proyeccion:
        return jsonify({"status": "error", "message": "Proyección no encontrada"}), 404
    return jsonify({"status": "success", "data": proyeccion})

@app.route('/api/proyecciones', methods=['POST'])
def api_save_proyeccion():
    data = request.get_json() or {}
    nombre_proyecto = data.get('nombre_proyecto', 'Proyección de Horas').strip()
    cliente = data.get('cliente', '').strip()
    modo_margen = data.get('modo_margen', 'costo')
    margen_global = data.get('margen_global', 15.0)
    unidad_escala = data.get('unidad_escala', 1000.0)
    items = data.get('items', [])
    proyeccion_id = data.get('id')
    
    if not items:
        return jsonify({"status": "error", "message": "Debe incluir al menos un profesional en la proyección"}), 400
        
    try:
        saved_id = database.save_proyeccion(
            nombre_proyecto=nombre_proyecto,
            cliente=cliente,
            modo_margen=modo_margen,
            margen_global=margen_global,
            unidad_escala=unidad_escala,
            items_data=items,
            proyeccion_id=proyeccion_id
        )
        u_email, u_nombre = get_current_user_info()
        database.log_activity(u_email, u_nombre, 'GUARDAR_PROYECCION', f'Guardó proyección ID {saved_id} ("{nombre_proyecto}") con {len(items)} profesionales', request.remote_addr)
        return jsonify({"status": "success", "id": saved_id, "message": "Proyección guardada correctamente"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error al guardar: {str(e)}"}), 500

@app.route('/api/proyecciones/<int:proyeccion_id>', methods=['DELETE'])
def api_delete_proyeccion(proyeccion_id):
    try:
        database.delete_proyeccion(proyeccion_id)
        return jsonify({"status": "success", "message": "Proyección eliminada"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# Export to CSV
@app.route('/api/export/csv', methods=['POST'])
def api_export_csv():
    data = request.get_json() or {}
    nombre_proyecto = data.get('nombre_proyecto', 'ProyectaBM_Proyeccion')
    items = data.get('items', [])
    modo_margen = data.get('modo_margen', 'costo')
    unidad_escala = float(data.get('unidad_escala', 1.0))
    
    u_email, u_nombre = get_current_user_info()
    database.log_activity(u_email, u_nombre, 'EXPORTAR_CSV', f'Exportó reporte de proyección ({len(items)} filas) a CSV', request.remote_addr)

    unit_str = "Normal ($)" if unidad_escala == 1 else ("Miles ($k)" if unidad_escala == 1000 else "Millones ($M)")
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow(['ProyectaBM - Reporte de Proyección de Horas'])
    writer.writerow(['Proyecto', nombre_proyecto])
    writer.writerow(['Modo de Margen', 'Sobre Costo (Markup)' if modo_margen == 'costo' else 'Sobre Venta (Profit Margin)'])
    writer.writerow([])
    writer.writerow([
        'Profesional',
        'Perfil de Costo',
        'Tarifa/Hora (UF/h)',
        'Horas Ingresadas',
        'Costo Directo (UF)',
        'Costo Directo ($ CLP)',
        'Margen Utilidad (%)',
        'Monto Utilidad (UF)',
        'Monto Utilidad ($ CLP)',
        'Precio Final Cliente (UF)',
        'Precio Final Cliente ($ CLP)'
    ])
    
    total_horas = 0
    total_costo_uf = 0
    total_costo_clp = 0
    total_utilidad_uf = 0
    total_utilidad_clp = 0
    total_precio_uf = 0
    total_precio_clp = 0
    
    for item in items:
        profesional = item.get('profesional', '')
        perfil = item.get('perfil_nombre', '')
        tarifa_uf = float(item.get('tarifa_costo', 0))
        horas = float(item.get('horas', 0))
        costo_uf = float(item.get('costo_total_uf', horas * tarifa_uf))
        costo_clp = float(item.get('costo_total', 0))
        margen = float(item.get('margen_porcentaje', 15))
        utilidad_uf = float(item.get('monto_utilidad_uf', 0))
        utilidad_clp = float(item.get('monto_utilidad', 0))
        precio_uf = float(item.get('precio_venta_uf', 0))
        precio_clp = float(item.get('precio_venta', 0))
        
        total_horas += horas
        total_costo_uf += costo_uf
        total_costo_clp += costo_clp
        total_utilidad_uf += utilidad_uf
        total_utilidad_clp += utilidad_clp
        total_precio_uf += precio_uf
        total_precio_clp += precio_clp
        
        writer.writerow([
            profesional,
            perfil,
            f"{tarifa_uf:.2f}",
            f"{horas:.1f}",
            f"{costo_uf:.2f}",
            f"{costo_clp:.0f}",
            f"{margen:.1f}%",
            f"{utilidad_uf:.2f}",
            f"{utilidad_clp:.0f}",
            f"{precio_uf:.2f}",
            f"{precio_clp:.0f}"
        ])
        
    writer.writerow([])
    writer.writerow([
        'TOTALES',
        '',
        '',
        f"{total_horas:.1f}",
        f"{total_costo_uf:.2f}",
        f"{total_costo_clp:.0f}",
        f"{(total_utilidad_clp / total_costo_clp * 100):.1f}%" if total_costo_clp > 0 else "0%",
        f"{total_utilidad_uf:.2f}",
        f"{total_utilidad_clp:.0f}",
        f"{total_precio_uf:.2f}",
        f"{total_precio_clp:.0f}"
    ])
    
    response = Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={"Content-Disposition": f"attachment;filename={nombre_proyecto.replace(' ', '_')}.csv"}
    )
    return response

if __name__ == '__main__':
    print("Iniciando ProyectaBM Web App en http://127.0.0.1:5000 ...")
    app.run(debug=True, host='0.0.0.0', port=5000)
